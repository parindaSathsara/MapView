import openpyxl, requests, json, time

API_KEY = "AIzaSyA0Z8_GzoG68yaI8pAEkA_03Ig1N-6qki8"
GEOCODE_URL = "https://maps.googleapis.com/maps/api/geocode/json"

wb = openpyxl.load_workbook('Vietnam Lat Long.xlsx')

hotels = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Reading sheet: {sheet_name} ({ws.max_row - 1} hotels)")
    for row in ws.iter_rows(min_row=2, values_only=True):
        city, name, cat, area, sic, address = row
        if not name or not address:
            continue
        # Clean address
        addr = str(address).replace('\n', ' ').strip()
        # Use city from the row for better geocoding
        city_str = str(city or sheet_name).strip()
        search_addr = f"{addr}, {city_str}, Vietnam"
        hotels.append({
            "city": city_str,
            "name": str(name),
            "cat": str(cat or ""),
            "area": str(area or ""),
            "sic": str(sic or ""),
            "address": addr,
            "search_address": search_addr
        })

print(f"Total hotels: {len(hotels)}")

results = []
for i, h in enumerate(hotels):
    params = {"address": h["search_address"], "key": API_KEY}
    resp = requests.get(GEOCODE_URL, params=params)
    data = resp.json()
    lat, lng = None, None
    if data["status"] == "OK" and data["results"]:
        loc = data["results"][0]["geometry"]["location"]
        lat = loc["lat"]
        lng = loc["lng"]
        print(f"[{i+1}/{len(hotels)}] OK: {h['name']} -> {lat}, {lng}")
    else:
        print(f"[{i+1}/{len(hotels)}] FAIL: {h['name']} ({data['status']})")
    
    results.append({
        "city": h["city"],
        "name": h["name"],
        "cat": h["cat"],
        "area": h["area"],
        "sic": h["sic"],
        "address": h["address"],
        "lat": lat,
        "lng": lng
    })
    time.sleep(0.1)  # Rate limiting

with open("hotels_geocoded.json", "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

ok = sum(1 for r in results if r["lat"] is not None)
print(f"\nGeocoded: {ok}/{len(results)}")
